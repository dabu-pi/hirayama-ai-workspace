"""
AI投資プロジェクト 管理シート xlsx テンプレート生成スクリプト
参照設計書: ai-invest/docs/sheet_design.md v1.3

実行方法:
  python ai-invest/templates/create_template.py

生成物:
  ai-invest/templates/AI投資管理シート_テンプレート.xlsx
  → Google Drive にアップロード後、「Googleスプレッドシートで開く」を選択
"""

from datetime import date
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

HEADER_FILL   = PatternFill("solid", fgColor="1A73E8")
HEADER_FONT   = Font(color="FFFFFF", bold=True, name="Meiryo", size=10)
HEADER_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_FONT     = Font(name="Meiryo", size=10)
BODY_ALIGN    = Alignment(vertical="center", wrap_text=False)

GREEN_FILL  = PatternFill("solid", fgColor="D9EAD3")
YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")
RED_FILL    = PatternFill("solid", fgColor="F4CCCC")
BLUE_FILL   = PatternFill("solid", fgColor="E8F0FE")

TODAY = date.today().isoformat()

UNIVERSE = [
    ("9843",  "ニトリHD",                "A"),
    ("3064",  "MonotaRO",               "A"),
    ("9432",  "日本電信電話（NTT）",     "A"),
    ("4684",  "オービック",              "B"),
    ("3769",  "GMOペイメントゲートウェイ","B"),
    ("4776",  "サイボウズ",              "B"),
    ("4307",  "野村総合研究所（NRI）",   "B"),
    ("6861",  "キーエンス",              "C"),
    ("6954",  "ファナック",              "C"),
    ("6273",  "SMC",                    "C"),
    ("6146",  "ディスコ",               "C"),
    ("6869",  "シスメックス",            "D"),
    ("2413",  "エムスリー",             "D"),
    ("4901",  "富士フイルムHD",          "D"),
    ("8766",  "東京海上HD",              "E"),
    ("8473",  "SBIホールディングス",     "E"),
]

def style_header(ws, row, headers, col_widths):
    for ci, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=row, column=ci, value=header)
        cell.fill   = HEADER_FILL
        cell.font   = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[row].height = 30
    ws.freeze_panes = ws.cell(row=2, column=1)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_body_style(ws, start_row, end_row, num_cols):
    for r in range(start_row, end_row + 1):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            if not cell.font or not cell.font.bold:
                cell.font = BODY_FONT
            cell.alignment = BODY_ALIGN
            cell.border = thin_border()

def add_dropdown(ws, col_letter, start_row, end_row, formula_str):
    """formula_str: '"選択肢1,選択肢2,..."' の形式（ダブルクォートで囲む）"""
    dv = DataValidation(
        type="list",
        formula1=formula_str,
        allow_blank=True,
        showDropDown=False,
    )
    dv.error      = "リストから選択してください"
    dv.errorTitle = "入力エラー"
    dv.prompt     = "リストから選択してください"
    ws.add_data_validation(dv)
    dv.sqref = f"{col_letter}{start_row}:{col_letter}{end_row}"


# ===== シート① 候補管理 =====

def create_sheet1(wb):
    ws = wb.create_sheet("①候補管理")

    headers = [
        "登録日", "コード", "銘柄名", "分類", "登録元",
        "登録理由", "AI根拠ソース", "一次コメント", "判定ステータス", "最終アクション"
    ]
    widths = [14, 9, 20, 8, 18, 25, 23, 25, 15, 15]
    style_header(ws, 1, headers, widths)

    # ドロップダウン
    add_dropdown(ws, "D", 2, 201, '"A,B,C,D,E,TBD"')
    add_dropdown(ws, "E", 2, 201, '"AI_UNIVERSE,AI_SCREENING,AI_NEWS,MANUAL_USER"')
    add_dropdown(ws, "G", 2, 201,
        '"UNIVERSE登録済み,SCREENING条件通過,決算材料,ニュース材料,ユーザー注目銘柄"')
    add_dropdown(ws, "I", 2, 201, '"未判定,判定済み"')
    add_dropdown(ws, "J", 2, 201, '"監視継続,保留,除外"')

    # UNIVERSE初期データ入力
    for i, (code, name, category) in enumerate(UNIVERSE, start=2):
        row_data = [
            TODAY, code, name, category, "AI_UNIVERSE",
            "UNIVERSE.md初期登録", "UNIVERSE登録済み", "", "未判定", "監視継続"
        ]
        for ci, val in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=ci, value=val)

    apply_body_style(ws, 1, 201, len(headers))
    return ws


# ===== シート② スクリーニング記録 =====

def create_sheet2(wb):
    ws = wb.create_sheet("②スクリーニング記録")

    headers = [
        "確認日", "コード", "銘柄名", "分類",
        "株価>25日MA", "株価>75日MA", "1か月騰落>+5%", "出来高OK",
        "売上成長OK", "営利成長OK", "PER≦50", "自己資本≧30%",
        "総合判定", "保留・除外理由", "監視メモ",
        "登録元", "データ確認日", "AIコメント"
    ]
    widths = [14, 9, 18, 8, 13, 13, 15, 11, 12, 11, 10, 14, 11, 25, 25, 18, 14, 25]
    style_header(ws, 1, headers, widths)

    # ドロップダウン
    add_dropdown(ws, "D", 2, 501, '"A,B,C,D,E"')
    for col_letter in ["E", "F", "G", "H", "I", "J", "K", "L"]:
        add_dropdown(ws, col_letter, 2, 501, '"○,✕"')
    add_dropdown(ws, "M", 2, 501, '"通過,保留,除外"')
    add_dropdown(ws, "P", 2, 501, '"AI_UNIVERSE,AI_SCREENING,AI_NEWS,MANUAL_USER"')

    apply_body_style(ws, 1, 501, len(headers))
    return ws


# ===== シート③ 取引記録 =====

def create_sheet3(wb):
    ws = wb.create_sheet("③取引記録")

    headers = [
        "日付", "種別", "コード", "銘柄名", "売買区分",
        "取引単価", "数量", "取引金額",
        "損切りライン", "部分利確ライン", "全利確ライン",
        "損益（円）", "損益（%）", "エグジット理由", "エントリー根拠・メモ"
    ]
    widths = [14, 11, 9, 18, 13, 11, 8, 13, 14, 16, 13, 13, 11, 14, 28]
    style_header(ws, 1, headers, widths)

    # ドロップダウン
    add_dropdown(ws, "B", 2, 501, '"ペーパー,実取引"')
    add_dropdown(ws, "E", 2, 501, '"買,売（部分）,売（全）"')
    add_dropdown(ws, "N", 2, 501, '"損切り,部分利確,全利確,期間利確,根拠消滅"')

    # H列: 取引金額の数式（=F*G）
    for row in range(2, 502):
        ws[f"H{row}"] = f"=IF(F{row}=\"\",\"\",F{row}*G{row})"

    apply_body_style(ws, 1, 501, len(headers))
    return ws


# ===== シート④ チャート型記録 =====

def create_sheet4(wb):
    ws = wb.create_sheet("④チャート型記録")

    headers = [
        "記録日", "コード", "銘柄名", "分類", "チャート型",
        "エントリー理由", "ルール通過状況", "エントリー価格",
        "損切りライン", "利確ルール", "結果", "損益率（%）",
        "失敗/成功理由", "振り返りメモ"
    ]
    widths = [14, 9, 18, 8, 26, 25, 15, 14, 13, 26, 9, 13, 25, 28]
    style_header(ws, 1, headers, widths)

    # ドロップダウン
    add_dropdown(ws, "D", 2, 501, '"A,B,C,D,E"')
    add_dropdown(ws, "E", 2, 501,
        '"25日線上の押し目,75日線上抜け後の初押し,高値更新ブレイク,'
        '出来高を伴うボックス上抜け,決算後ギャップアップ継続,'
        '25日線割れ反発失敗,その他"'
    )
    add_dropdown(ws, "G", 2, 501, '"通過,一部未確認,特例"')
    add_dropdown(ws, "J", 2, 501, '"標準（+10%部分/+20%全）,その他"')
    add_dropdown(ws, "K", 2, 501, '"勝ち,負け,保留"')

    # I列: 損切りライン = エントリー価格 × 0.95
    for row in range(2, 502):
        ws[f"I{row}"] = f"=IF(H{row}=\"\",\"\",ROUND(H{row}*0.95,0))"

    apply_body_style(ws, 1, 501, len(headers))
    return ws


# ===== メイン =====

def main():
    wb = Workbook()
    # デフォルトシートを削除
    wb.remove(wb.active)

    create_sheet1(wb)
    create_sheet2(wb)
    create_sheet3(wb)
    create_sheet4(wb)

    output_path = "ai-invest/templates/AI投資管理シート_テンプレート.xlsx"
    wb.save(output_path)
    print(f"[OK] 生成完了: {output_path}")
    print("Google Drive にアップロードし、Googleスプレッドシートで開いて使用してください。")


if __name__ == "__main__":
    main()
