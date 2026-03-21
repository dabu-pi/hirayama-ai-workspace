"""
テンプレートと出力ファイルの merged_cells を比較するスクリプト。
対象エリア:
  - 性別:    rows 20-25, cols AL-AO
  - 保険種別: rows 6-14,  cols CB-CM (および隣接 BZ-CA, CN-CS)
  - 単独区分: rows 6-14,  cols CT-CY (および隣接 CZ-DA)
  - 本家区分: rows 6-14,  cols DB-DM (および隣接 DN-DO)
"""
import sys
import glob
import os
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

TEMPLATE = "application_template.xlsx"
SHEET    = "新　様式第5号"

wb_tmpl = load_workbook(TEMPLATE, data_only=True)
ws_tmpl = wb_tmpl[SHEET]

# ---- 対象エリア絞り込み ----
# col: AL(38)-AO(41), BZ(78)-DO(119), rows 6-26
def is_target(mr):
    row_ok = 6 <= mr.min_row <= 26
    col_ok = (
        (38 <= mr.min_col <= 41) or   # 性別
        (78 <= mr.min_col <= 119)      # 保険種別〜本家区分
    )
    return row_ok and col_ok

print("=== [TEMPLATE] 対象エリアの merged_cells ===")
tmpl_ranges = [mr for mr in ws_tmpl.merged_cells.ranges if is_target(mr)]
for mr in sorted(tmpl_ranges, key=lambda r: (r.min_row, r.min_col)):
    print(f"  {str(mr):20s}  rows={mr.min_row}-{mr.max_row}  cols={mr.min_col}-{mr.max_col}")

# ---- 出力ファイルを探す ----
search_dirs = ["output/2026-03", "output", "."]
out_file = None
for d in search_dirs:
    files = sorted(glob.glob(f"{d}/*.xlsx"), key=os.path.getmtime, reverse=True)
    if files:
        out_file = files[0]
        break

if not out_file:
    print("\n出力ファイルが見つかりません。テンプレートのみ表示します。")
    sys.exit(0)

print(f"\n=== [OUTPUT: {out_file}] 対象エリアの merged_cells ===")
wb_out = load_workbook(out_file, data_only=True)
# シート名が一致しない可能性があるので active を使う
ws_out_candidates = [ws_tmpl.title]
if SHEET in wb_out.sheetnames:
    ws_out = wb_out[SHEET]
else:
    ws_out = wb_out.active

out_ranges = [mr for mr in ws_out.merged_cells.ranges if is_target(mr)]
for mr in sorted(out_ranges, key=lambda r: (r.min_row, r.min_col)):
    print(f"  {str(mr):20s}  rows={mr.min_row}-{mr.max_row}  cols={mr.min_col}-{mr.max_col}")

# ---- 差分 ----
tmpl_set = {str(mr) for mr in tmpl_ranges}
out_set  = {str(mr) for mr in out_ranges}

print("\n=== テンプレートにあって出力にない (消えた結合) ===")
lost = sorted(tmpl_set - out_set)
if lost:
    for r in lost:
        print(f"  LOST : {r}")
else:
    print("  なし")

print("\n=== 出力にあってテンプレートにない (追加された結合) ===")
added = sorted(out_set - tmpl_set)
if added:
    for r in added:
        print(f"  ADDED: {r}")
else:
    print("  なし")

# ---- SELECTION_SPLIT_MAP の unmerge 対象が実際に消えているか確認 ----
print("\n=== SELECTION_SPLIT_MAP の unmerge 対象 → 出力での状態 ===")
SPLIT_MAP_FULLS = [
    ("gender_男",  "AL21:AO22"),
    ("gender_女",  "AL23:AO24"),
    ("ins_1",      "CB8:CE10"),
    ("ins_2",      "CF8:CI10"),
    ("ins_3",      "CJ8:CM10"),
    ("ins_4",      "CB11:CE13"),
    ("ins_5",      "CF11:CI13"),
    ("ins_6",      "CJ11:CM13"),
    ("tankei_1",   "CT8:CY9"),
    ("tankei_2",   "CT10:CY11"),
    ("tankei_3",   "CT12:CY13"),
    ("honke_DB8",  "DB8:DG9"),
    ("honke_DB10", "DB10:DG11"),
    ("honke_DB12", "DB12:DG13"),
    ("honke_DH8",  "DH8:DM9"),
    ("honke_DH12", "DH12:DM13"),
]
for key, full in SPLIT_MAP_FULLS:
    in_tmpl = full in tmpl_set
    in_out  = full in out_set
    status  = "OK(消えた)" if (in_tmpl and not in_out) else ("残存!" if in_out else "テンプレに無し")
    print(f"  {key:15s}  {full:15s}  tmpl={in_tmpl}  out={in_out}  => {status}")

# ---- 代替として追加されるべき sub-range が存在するか ----
print("\n=== 分割後に存在すべき label/marker merge → 出力での状態 ===")
SPLIT_MAP_SUBS = [
    ("gender_男 label",  "AL21:AO21"),
    ("gender_男 marker", "AL22:AO22"),
    ("gender_女 label",  "AL23:AO23"),
    ("gender_女 marker", "AL24:AO24"),
    ("ins_1 label",      "CB8:CE9"),
    ("ins_1 marker",     "CB10:CE10"),
    ("tankei_1 label",   "CT8:CY8"),
    ("tankei_1 marker",  "CT9:CY9"),
    ("honke_DB8 label",  "DB8:DG8"),
    ("honke_DB8 marker", "DB9:DG9"),
    ("honke_DH8 label",  "DH8:DM8"),
    ("honke_DH8 marker", "DH9:DM9"),
]
for name, rng in SPLIT_MAP_SUBS:
    in_out = rng in out_set
    print(f"  {name:20s}  {rng:15s}  out={in_out}  => {'存在' if in_out else '欠落!'}")
